# coding=UTF-8
import requests
import re
from openpyxl import workbook  
from openpyxl import load_workbook
from xlutils.copy import copy
import xlwt
import xlrd
import requests
import json
import re
import csv
import urllib3
import time
import types
import FilterData
import pandas

MERGED_URL = 'http://136.17.76.164:8082/changes/?q=status:merged&n=25&O=81'
global val
val = 1
def dict_get(dict, objkey, default):
	tmp = dict
	for k,v in tmp.items():
		if k == objkey:
			return v
		else:
			if type(v) is types.DictType:
				ret = dict_get(v, objkey, default)
				if ret is not default:
					return ret
#	return v
def time_cmp(first_time, second_time):
    print(first_time)
    print(second_time)
    return int(first_time) - int(second_time)

def time_format(time):
    time = time.replace('.000000000', "")
    time = time.replace(' ', "")
    time = time.replace(':', "")
    time = time.replace('-', "")
    return time
	
def requesst(url,limit_time,start):
	continue_flag = 1
	global val
    # while(url):
	headers = {
		'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.131 Safari/537.36',
		'Cookie': 'GERRIT_UI=GWT; GerritAccount=aSseprswLuBU0N6Zr8Qo6K.rLgUxknKwCq; XSRF_TOKEN=aSseprsTSA4MCp0r1Gz3BJQQGtJqnZ9wUa'
	}

	request_url = url + "&S=" + str(start)
	data = requests.get(request_url, headers=headers).text
    #去掉多余的几个垃圾字符
	remove = re.compile('\)\]\}\'')
	data = re.sub(remove, "", data)
	data_json = json.loads(data)
	print data
#	workbook = xlwt.Workbook(encoding='utf-8')
#	worksheet = workbook.add_sheet('sheet1')
#	worksheet.write(0, 0, label='project')
#	worksheet.write(0, 1, label='branch')
#	worksheet.write(0, 2, label='subject')
#	worksheet.write(0, 3, label='owner')
#	worksheet.write(0, 4, label='updated')
	workbook = xlrd.open_workbook('test.xls')
	wfile = copy(workbook)
	wsheet = wfile.get_sheet('sheet1')
	val1 = val
	val2 = val
	val3 = val
	val4 = val
	val5 = val
	for list_item in data_json:
		for key,value in list_item.items():
			if key =="project":
				wsheet.write(val1, 0, value)
				val1 += 1
			elif key == "branch":
				wsheet.write(val2, 1, value)
				val2 += 1
			elif key == "subject":
				wsheet.write(val3, 2, value)
				val3 += 1
			elif key == "owner":
				v = dict_get(value, 'name', None)
				wsheet.write(val4, 3, v)
				val4 += 1
			elif key == "updated":
				vvalue = time_format(value)
				if (time_cmp(vvalue, limit_time) < 0):
					continue_flag = 0
					break
				else:
					wsheet.write(val5, 4, value)
					val5 += 1
			else:
				pass
	wfile.save('test.xls')
	val = val5
	return continue_flag, start+len(data_json), val

def main ():
	limit_time = input("format:20190418000000\n")
	url = MERGED_URL
	continue_flag = 1
	start = 0	
#	global val
	workbook = xlwt.Workbook(encoding='utf-8')
	worksheet = workbook.add_sheet('sheet1')
	worksheet.write(0, 0, label='project')
	worksheet.write(0, 1, label='branch')
	worksheet.write(0, 2, label='subject')
	worksheet.write(0, 3, label='owner')
	worksheet.write(0, 4, label='updated')
	workbook.save('test.xls')
	while (continue_flag):
		continue_flag, start, val = requesst(url,limit_time,start)
		val +=1
if __name__ == '__main__':
	main()
	#参数修改对应关系（excel表名，仓库关键字，仓库分支名）
	FilterData.filder_project('test.xls','xxx','master')
	
	
	
	
	
